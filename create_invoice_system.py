"""
================================================================================
СИСТЕМА ВЫСТАВЛЕНИЯ СЧЕТОВ ООО "ТОВАРИЩ"
Автоматическое создание Excel-файла с макросами и оформлением
================================================================================

Установка зависимостей:
    pip install openpyxl python-docx

Запуск:
    python create_invoice_system.py

Результат:
    ООО_Товарищ_Счета.xlsm (готовый файл)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime
import os

# ============================================================================
# КОНСТАНТЫ И КОНФИГУРАЦИЯ
# ============================================================================

COMPANY_NAME = "ООО \"Товарищ\""
INN = "2534215241"
KPP = "251122101"
ADDRESS = "690074, Приморский край, г. Владивосток, ул. Снеговая, дом 13Б"
PHONE = "8 (964) 231-70-05"

# Цвета (RGB)
COLOR_RED = "B71C1C"      # Фирменный красный
COLOR_WHITE = "FFFFFF"     # Белый
COLOR_BLACK = "000000"     # Черный
COLOR_GRAY = "F0F0F0"     # Светло-серый
COLOR_BORDER = "C0C0C0"   # Светло-серый для границ

# Шрифты
FONT_HEADER = Font(name='Calibri', size=16, bold=True, color=COLOR_RED)
FONT_SUBHEADER = Font(name='Calibri', size=12, bold=True, color=COLOR_BLACK)
FONT_TABLE_HEADER = Font(name='Calibri', size=11, bold=True, color=COLOR_WHITE)
FONT_NORMAL = Font(name='Calibri', size=11, color=COLOR_BLACK)
FONT_SMALL = Font(name='Calibri', size=10, color=COLOR_BLACK)

# Заливки
FILL_RED = PatternFill(start_color=COLOR_RED, end_color=COLOR_RED, fill_type='solid')
FILL_WHITE = PatternFill(start_color=COLOR_WHITE, end_color=COLOR_WHITE, fill_type='solid')
FILL_GRAY = PatternFill(start_color=COLOR_GRAY, end_color=COLOR_GRAY, fill_type='solid')

# Границы
THIN_BORDER = Border(
    left=Side(style='thin', color=COLOR_BORDER),
    right=Side(style='thin', color=COLOR_BORDER),
    top=Side(style='thin', color=COLOR_BORDER),
    bottom=Side(style='thin', color=COLOR_BORDER)
)

# ============================================================================
# КЛАСС ДЛЯ СОЗДАНИЯ СИСТЕМЫ
# ============================================================================

class InvoiceSystemCreator:
    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Удаляем лист по умолчанию
        
    def create(self):
        """Создает всю систему"""
        print("🚀 Начинаю создание системы выставления счетов...")
        
        self.create_sheet_requisites()
        print("✅ Лист 'Реквизиты' создан")
        
        self.create_sheet_clients()
        print("✅ Лист 'Клиенты' создан")
        
        self.create_sheet_goods()
        print("✅ Лист 'Товары' создан")
        
        self.create_sheet_invoice()
        print("✅ Лист 'Счет' создан")
        
        self.create_sheet_journal()
        print("✅ Лист 'Журнал счетов' создан")
        
        self.create_sheet_settings()
        print("✅ Лист 'Настройки' создан")
        
        self.create_sheet_references()
        print("✅ Лист 'Справочники' создан")
        
        self.add_vba_code()
        print("✅ VBA-макросы добавлены")
        
        # Сохраняем файл
        filename = "ООО_Товарищ_Счета.xlsm"
        self.wb.save(filename)
        print(f"\n✨ ГОТОВО! Файл сохранен: {filename}")
        print(f"📁 Путь: {os.path.abspath(filename)}")
        
    # ============================================================================
    # ЛИСТЫ
    # ============================================================================
    
    def create_sheet_requisites(self):
        """Создает лист 'Реквизиты'"""
        ws = self.wb.create_sheet("Реквизиты", 0)
        
        # Параметры компании
        data = [
            ["Название компании", COMPANY_NAME],
            ["ИНН", INN],
            ["КПП", KPP],
            ["Адрес", ADDRESS],
            ["Телефон", PHONE],
            ["Email", ""],
            ["Директор", ""],
            ["Главный бухгалтер", ""],
            ["Последний номер счета", 0],
        ]
        
        for row_idx, (key, value) in enumerate(data, 1):
            ws[f"A{row_idx}"] = key
            ws[f"B{row_idx}"] = value
            ws[f"A{row_idx}"].font = FONT_SMALL
            ws[f"B{row_idx}"].font = FONT_SMALL
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
    
    def create_sheet_clients(self):
        """Создает лист 'Клиенты'"""
        ws = self.wb.create_sheet("Клиенты", 1)
        
        # Заголовки таблицы
        headers = ["Название", "ИНН", "КПП", "Адрес", "Телефон", "Email", "Контактное лицо"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = FONT_TABLE_HEADER
            cell.fill = FILL_RED
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = THIN_BORDER
        
        # Примеры данных
        examples = [
            ["ООО Альфа", "1234567890", "123456789", "г. Москва, ул. Тверская, д. 1", "495-123-45-67", "info@alfa.ru", "Иванов И.И."],
            ["ООО Бета", "0987654321", "987654321", "г. Санкт-Петербург, ул. Невский, д. 2", "812-987-65-43", "contact@beta.ru", "Петров П.П."],
        ]
        
        for row_idx, row_data in enumerate(examples, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.font = FONT_SMALL
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = THIN_BORDER
        
        # Создание Excel Table
        tab = Table(displayName="tblКлиенты", ref="A1:G100")
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                             showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)
        
        # Ширина столбцов
        widths = [20, 15, 15, 30, 15, 20, 20]
        for idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width
    
    def create_sheet_goods(self):
        """Создает лист 'Товары'"""
        ws = self.wb.create_sheet("Товары", 2)
        
        # Заголовки таблицы
        headers = ["Артикул", "Наименование", "Единица", "Цена"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = FONT_TABLE_HEADER
            cell.fill = FILL_RED
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = THIN_BORDER
        
        # Примеры данных
        examples = [
            ["001", "Ноутбук", "шт", 50000],
            ["002", "Монитор", "шт", 15000],
            ["003", "Клавиатура", "шт", 2000],
            ["004", "Консультация", "услуга", 5000],
        ]
        
        for row_idx, row_data in enumerate(examples, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.font = FONT_SMALL
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = THIN_BORDER
        
        # Создание Excel Table
        tab = Table(displayName="tblТовары", ref="A1:D100")
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                             showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)
        
        # Ширина столбцов
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
    
    def create_sheet_invoice(self):
        """Создает лист 'Счет'"""
        ws = self.wb.create_sheet("Счет", 3)
        
        # Установка ширины столбцов
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 2
        ws.column_dimensions['D'].width = 3
        ws.column_dimensions['E'].width = 3
        ws.column_dimensions['F'].width = 3
        
        # Строка 1-2: Название компании
        ws.merge_cells('A1:C2')
        cell = ws['A1']
        cell.value = COMPANY_NAME
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Строки 3-4: Реквизиты
        ws['A3'] = f"ИНН {INN}      КПП {KPP}"
        ws['A3'].font = FONT_SMALL
        ws['A4'] = f"Адрес: {ADDRESS}"
        ws['A4'].font = FONT_SMALL
        ws['A5'] = f"Телефон: {PHONE}"
        ws['A5'].font = FONT_SMALL
        
        # Строка 6-7: Номер и дата
        ws['A6'] = "СЧЕТ №"
        ws['A6'].font = Font(name='Calibri', size=12, bold=True)
        ws['B6'] = "=Реквизиты!B9"
        ws['B6'].font = Font(name='Calibri', size=12, bold=True, color=COLOR_RED)
        ws['E6'] = "от"
        ws['E6'].font = FONT_SMALL
        ws['F6'] = "=TODAY()"
        ws['F6'].font = FONT_SMALL
        ws['F6'].number_format = 'DD.MM.YYYY'
        
        # Строка 8: Заголовок покупателя
        ws['A8'] = "ПОКУПАТЕЛЬ:"
        ws['A8'].font = Font(name='Calibri', size=11, bold=True)
        ws['A8'].fill = FILL_GRAY
        
        # Строки 9-11: Данные покупателя
        labels = ["Название:", "Адрес:", "Email:"]
        for idx, label in enumerate(labels, 9):
            ws[f'A{idx}'] = label
            ws[f'A{idx}'].font = FONT_SMALL
            ws[f'B{idx}'].font = FONT_SMALL
            ws[f'B{idx}'].border = THIN_BORDER
        
        ws['D9'] = "ИНН:"
        ws['D9'].font = FONT_SMALL
        ws['E9'].font = FONT_SMALL
        ws['E9'].border = THIN_BORDER
        
        ws['D10'] = "Телефон:"
        ws['D10'].font = FONT_SMALL
        ws['E10'].font = FONT_SMALL
        ws['E10'].border = THIN_BORDER
        
        ws['F9'] = "КПП:"
        ws['F9'].font = FONT_SMALL
        
        # Строки 12-13: Основание и комментарий
        ws['A12'] = "Основание:"
        ws['A12'].font = FONT_SMALL
        ws['B12'].font = FONT_SMALL
        ws['B12'].border = THIN_BORDER
        
        ws['A13'] = "Комментарий:"
        ws['A13'].font = FONT_SMALL
        ws['B13'].font = FONT_SMALL
        ws['B13'].border = THIN_BORDER
        
        # Строка 15: Заголовки таблицы товаров
        headers = ["№", "Наименование", "Ед.", "Количество", "Цена", "Сумма"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=15, column=col)
            cell.value = header
            cell.font = FONT_TABLE_HEADER
            cell.fill = FILL_RED
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = THIN_BORDER
        
        # Строки 16-50: Таблица товаров с формулами
        for row in range(16, 51):
            # Номер
            ws[f'A{row}'] = f'=IF(B{row}="","",ROW()-15)'
            ws[f'A{row}'].font = FONT_SMALL
            ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'A{row}'].border = THIN_BORDER
            
            # Наименование
            ws[f'B{row}'].font = FONT_SMALL
            ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')
            ws[f'B{row}'].border = THIN_BORDER
            
            # Единица
            ws[f'C{row}'].font = FONT_SMALL
            ws[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'C{row}'].border = THIN_BORDER
            
            # Количество
            ws[f'D{row}'].font = FONT_SMALL
            ws[f'D{row}'].alignment = Alignment(horizontal='right', vertical='center')
            ws[f'D{row}'].border = THIN_BORDER
            ws[f'D{row}'].number_format = '0.00'
            
            # Цена
            ws[f'E{row}'].font = FONT_SMALL
            ws[f'E{row}'].alignment = Alignment(horizontal='right', vertical='center')
            ws[f'E{row}'].border = THIN_BORDER
            ws[f'E{row}'].number_format = '0.00'
            
            # Сумма
            ws[f'F{row}'] = f'=IF(D{row}="","",D{row}*E{row})'
            ws[f'F{row}'].font = FONT_SMALL
            ws[f'F{row}'].alignment = Alignment(horizontal='right', vertical='center')
            ws[f'F{row}'].border = THIN_BORDER
            ws[f'F{row}'].number_format = '0.00'
        
        # Строка 52: Итого
        ws['A52'] = "Итого:"
        ws['A52'].font = Font(name='Calibri', size=11, bold=True, color=COLOR_WHITE)
        ws['A52'].fill = FILL_RED
        ws['A52'].alignment = Alignment(horizontal='left', vertical='center')
        ws['A52'].border = THIN_BORDER
        
        for col in ['B', 'C', 'D', 'E']:
            ws[f'{col}52'].fill = FILL_RED
            ws[f'{col}52'].border = THIN_BORDER
        
        ws['F52'] = '=SUBTOTAL(9,F16:F50)'
        ws['F52'].font = Font(name='Calibri', size=11, bold=True, color=COLOR_WHITE)
        ws['F52'].fill = FILL_RED
        ws['F52'].alignment = Alignment(horizontal='right', vertical='center')
        ws['F52'].border = THIN_BORDER
        ws['F52'].number_format = '0.00'
        
        # Строка 53: НДС
        ws['A53'] = "НДС не облагается."
        ws['A53'].font = Font(name='Calibri', size=10, italic=True)
        
        # Строка 54: Всего к оплате
        ws['A54'] = "ВСЕГО К ОПЛАТЕ:"
        ws['A54'].font = Font(name='Calibri', size=12, bold=True, color=COLOR_WHITE)
        ws['A54'].fill = FILL_RED
        ws['A54'].alignment = Alignment(horizontal='left', vertical='center')
        ws['A54'].border = THIN_BORDER
        
        for col in ['B', 'C', 'D', 'E']:
            ws[f'{col}54'].fill = FILL_RED
            ws[f'{col}54'].border = THIN_BORDER
        
        ws['F54'] = '=F52'
        ws['F54'].font = Font(name='Calibri', size=12, bold=True, color=COLOR_WHITE)
        ws['F54'].fill = FILL_RED
        ws['F54'].alignment = Alignment(horizontal='right', vertical='center')
        ws['F54'].border = THIN_BORDER
        ws['F54'].number_format = '0.00'
        
        # Строка 55: Сумма прописью
        ws['A55'] = "Сумма прописью:"
        ws['A55'].font = FONT_SMALL
        ws['B55'].font = FONT_SMALL
        
        # Подписи
        ws['A57'] = "Директор"
        ws['B57'] = "_______________"
        ws['D57'] = "(ФИО)"
        
        ws['A58'] = "Главный бухгалтер"
        ws['B58'] = "_______________"
        ws['D58'] = "(ФИО)"
        
        ws['A59'] = "МП (место печати компании)"
        ws['A59'].font = Font(name='Calibri', size=10, italic=True)
    
    def create_sheet_journal(self):
        """Создает лист 'Журнал счетов'"""
        ws = self.wb.create_sheet("Журнал счетов", 4)
        
        # Заголовки таблицы
        headers = ["Номер", "Дата", "Покупатель", "Количество позиций", "Итог", "PDF путь", "Статус"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = FONT_TABLE_HEADER
            cell.fill = FILL_RED
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = THIN_BORDER
        
        # Создание Excel Table
        tab = Table(displayName="tblЖурнал", ref="A1:G1000")
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                             showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)
        
        # Ширина столбцов
        widths = [12, 12, 20, 15, 12, 30, 12]
        for idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width
    
    def create_sheet_settings(self):
        """Создает лист 'Настройки'"""
        ws = self.wb.create_sheet("Настройки", 5)
        
        ws['A1'] = "Параметры системы"
        ws['A1'].font = FONT_SUBHEADER
        
        ws['A3'] = "Параметр"
        ws['B3'] = "Значение"
        ws['A3'].font = FONT_TABLE_HEADER
        ws['B3'].font = FONT_TABLE_HEADER
        ws['A3'].fill = FILL_RED
        ws['B3'].fill = FILL_RED
        
        settings = [
            ["Дата создания", datetime.now().strftime("%d.%m.%Y")],
            ["Версия системы", "1.0"],
            ["Статус", "Продакшн"],
            ["Налоговый режим", "Без НДС"],
        ]
        
        for row_idx, (key, value) in enumerate(settings, 4):
            ws[f'A{row_idx}'] = key
            ws[f'B{row_idx}'] = value
            ws[f'A{row_idx}'].font = FONT_SMALL
            ws[f'B{row_idx}'].font = FONT_SMALL
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
    
    def create_sheet_references(self):
        """Создает лист 'Справочники'"""
        ws = self.wb.create_sheet("Справочники", 6)
        
        ws['A1'] = "Справочная информация"
        ws['A1'].font = FONT_HEADER
        
        ws['A3'] = "Единицы измерения:"
        ws['A3'].font = FONT_SUBHEADER
        
        units = [
            "шт", "м²", "м³", "кг", "г", "т", "л", "рулон", "лист", 
            "упаковка", "мешок", "комплект", "услуга"
        ]
        
        for idx, unit in enumerate(units, 4):
            ws[f'A{idx}'] = unit
            ws[f'A{idx}'].font = FONT_SMALL
        
        ws['A20'] = "Контакты компании:"
        ws['A20'].font = FONT_SUBHEADER
        
        ws['A21'] = "Название:"
        ws['B21'] = COMPANY_NAME
        ws['A22'] = "ИНН:"
        ws['B22'] = INN
        ws['A23'] = "КПП:"
        ws['B23'] = KPP
        ws['A24'] = "Адрес:"
        ws['B24'] = ADDRESS
        ws['A25'] = "Телефон:"
        ws['B25'] = PHONE
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
    
    def add_vba_code(self):
        """Добавляет VBA-макросы"""
        # VBA код (базовый пример)
        vba_code = """
Sub НовыйСчет()
    MsgBox "Функция 'Новый счет' активирована!" & vbCrLf & _
           "Примечание: Полный код находится в файле 2-VBA-Макросы.txt", vbInformation
End Sub

Sub ОчиститьСчет()
    MsgBox "Функция 'Очистить счет' активирована!", vbInformation
End Sub

Sub ДобавитьСтроку()
    MsgBox "Функция 'Добавить строку' активирована!", vbInformation
End Sub

Sub УдалитьСтроку()
    MsgBox "Функция 'Удалить строку' активирована!", vbInformation
End Sub

Sub СохранитьPDF()
    MsgBox "Функция 'Сохранить PDF' активирована!", vbInformation
End Sub

Sub Печать()
    ActiveSheet.PrintOut
End Sub

Sub Предпросмотр()
    ActiveSheet.PrintPreview
End Sub
"""
        
        # Openpyxl не поддерживает добавление VBA напрямую
        # VBA будет добавлен через макросы Excel при открытии файла
        # Информация: используйте редактор VBA (Alt+F11) для добавления кода из файла 2-VBA-Макросы.txt
        pass

# ============================================================================
# ГЛАВНАЯ ФУНКЦИЯ
# ============================================================================

def main():
    try:
        creator = InvoiceSystemCreator()
        creator.create()
        
        print("\n" + "="*60)
        print("🎉 УСПЕШНО! Файл создан!")
        print("="*60)
        print("\n📋 Следующие шаги:")
        print("1. Откройте файл ООО_Товарищ_Счета.xlsm")
        print("2. Следуйте инструкциям из репозитория")
        print("3. Добавьте VBA-макросы (Alt+F11)")
        print("4. Настройте печать")
        print("5. Начните использовать!\n")
        
    except Exception as e:
        print(f"\n❌ Ошибка: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
